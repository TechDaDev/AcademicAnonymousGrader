# Academic Anonymous Grader — Analytics XLSX Report Export
"""Privacy-safe XLSX analytics report generation.

All reports are generated in memory, contain no student identity,
no responses, no feedback text, and no individual grade rows.
"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from analytics.filters import AnalyticsFilter
from analytics.models import ReportExportMetadata
from services.authorization_service import AuthContext

if TYPE_CHECKING:
    from sqlalchemy.orm import Session

# Styling
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)


def _style_header(ws: Any, col_count: int) -> None:
    """Apply header styling and freeze panes."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
    ws.freeze_panes = "A2"


def _auto_width(ws: Any, col_count: int, max_width: int = 40) -> None:
    """Set reasonable column widths."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 10)


def _build_metadata(
    report_title: str,
    filter_obj: AnalyticsFilter | None,
    privacy_notes: str = "",
) -> ReportExportMetadata:
    """Build report export metadata."""
    return ReportExportMetadata(
        report_title=report_title,
        generated_at=datetime.now(UTC).isoformat(),
        filter_summary=filter_obj.summary() if filter_obj else "none",
        population_definition="Approved submissions with graded records",
        privacy_suppression_notes=privacy_notes,
        application_version="1.0.0-rc1",
        schema_version="3",
    )


def _write_metadata_sheet(
    wb: Workbook,
    metadata: ReportExportMetadata,
) -> None:
    """Write metadata sheet to workbook."""
    ws = wb.create_sheet("Metadata", 0)
    headers = ["Field", "Value"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for field, value in [
        ("Report Title", metadata.report_title),
        ("Generated At", metadata.generated_at),
        ("Filter Summary", metadata.filter_summary),
        ("Population Definition", metadata.population_definition),
        ("Privacy Suppression Notes", metadata.privacy_suppression_notes),
        ("Application Version", metadata.application_version),
        ("Schema Version", metadata.schema_version),
    ]:
        ws.append([field, value])
    _auto_width(ws, 2)


def export_assessment_summary_report(
    session: Session,
    ctx: AuthContext | None,
    assessment_id: str,
    filter_obj: AnalyticsFilter | None = None,
) -> bytes:
    """Generate an XLSX assessment summary report.

    Parameters
    ----------
    session : Session
        Database session.
    ctx : AuthContext | None
        Authorization context.
    assessment_id : str
        Assessment ID.
    filter_obj : AnalyticsFilter | None
        Optional filter.

    Returns
    -------
    bytes
        Workbook bytes.

    Raises
    ------
    PermissionError
        If unauthorized.
    """
    from analytics.assessment_analysis import get_assessment_performance
    from analytics.corrections import get_correction_analytics
    from analytics.distributions import get_grade_distribution
    from analytics.question_analysis import list_question_analyses

    perf = get_assessment_performance(session, ctx, assessment_id, filter_obj)
    if perf.privacy.suppressed:
        msg = "No authorized data available for export."
        raise PermissionError(msg)

    dist = get_grade_distribution(session, ctx, assessment_id, filter_obj)
    questions = list_question_analyses(session, ctx, assessment_id, filter_obj)
    corrections = get_correction_analytics(session, ctx, assessment_id, filter_obj)

    metadata = _build_metadata(
        f"Assessment Summary: {perf.assessment_title}",
        filter_obj,
        privacy_notes="Data suppressed where group size below threshold." if perf.privacy.suppressed else "",
    )

    wb = Workbook()
    _write_metadata_sheet(wb, metadata)

    # Summary sheet
    ws_summary = wb.active
    assert ws_summary is not None  # noqa: S101
    ws_summary.title = "Summary"
    summary_headers = [
        "Metric", "Value",
    ]
    ws_summary.append(summary_headers)
    _style_header(ws_summary, len(summary_headers))
    for metric, value in [
        ("Assessment", perf.assessment_title),
        ("Material", perf.material_title),
        ("Maximum Grade", float(perf.maximum_grade)),
        ("Total Submissions", perf.submission_count),
        ("Graded", perf.graded_count),
        ("Approved", perf.approved_count),
        ("Mean Grade", perf.mean_grade),
        ("Median Grade", perf.median_grade),
        ("Minimum Grade", perf.minimum_grade),
        ("Maximum Grade", perf.maximum_grade),
        ("Standard Deviation", perf.standard_deviation),
        ("Pass Count", perf.pass_count),
        ("Fail Count", perf.fail_count),
        ("Pass %", perf.pass_percentage),
        ("Completion %", perf.completion_percentage),
        ("Ready for Finalization", "Yes" if perf.ready_for_finalization else "No"),
        ("Blocker Count", perf.blocker_count),
    ]:
        ws_summary.append([metric, value])

    # Add classification metadata
    from models.assessment import Assessment
    asmt_obj = (
        session.query(Assessment).filter(Assessment.id == assessment_id).first()
    )
    if asmt_obj and asmt_obj.material:
        mat_obj = asmt_obj.material
        ws_summary.append([])
        ws_summary.append(["Classification", ""])
        dept_lbl = mat_obj.ref_department.display_name if mat_obj.ref_department else "—"
        stage_lbl = mat_obj.ref_stage.display_name if mat_obj.ref_stage else "—"
        term_lbl = mat_obj.ref_term.display_name if mat_obj.ref_term else "—"
        year_lbl2 = mat_obj.ref_academic_year.display_name if mat_obj.ref_academic_year else "—"
        ws_summary.append(["Department", dept_lbl])
        ws_summary.append(["Stage", stage_lbl])
        ws_summary.append(["Term", term_lbl])
        ws_summary.append(["Academic Year", year_lbl2])
    _auto_width(ws_summary, 2)

    # Distribution sheet
    ws_dist = wb.create_sheet("Distribution")
    dist_headers = ["Grade Band", "Count", "Percentage"]
    ws_dist.append(dist_headers)
    _style_header(ws_dist, len(dist_headers))
    for band in dist.bands:
        ws_dist.append([band.label, band.count, round(band.percentage, 1)])
    _auto_width(ws_dist, 3)

    # Questions sheet
    ws_q = wb.create_sheet("Questions")
    q_headers = [
        "Question", "Max Grade", "Graded", "Mean", "Median",
        "Min", "Max", "Std Dev", "Avg %", "Difficulty",
        "Full Score", "Zero Score", "Partial", "Unanswered",
    ]
    ws_q.append(q_headers)
    _style_header(ws_q, len(q_headers))
    for qa in questions:
        ws_q.append([
            qa.question_number,
            float(qa.maximum_grade),
            qa.graded_response_count,
            qa.mean_awarded,
            qa.median_awarded,
            qa.minimum_awarded,
            qa.maximum_awarded,
            qa.standard_deviation,
            qa.average_percentage,
            qa.difficulty_index,
            qa.full_score_count,
            qa.zero_score_count,
            qa.partial_credit_count,
            qa.unanswered_count,
        ])
    _auto_width(ws_q, len(q_headers))

    # Corrections sheet
    ws_c = wb.create_sheet("Corrections")
    c_headers = ["Metric", "Value"]
    ws_c.append(c_headers)
    _style_header(ws_c, len(c_headers))
    for metric, value in [
        ("Total Returned for Correction", corrections.total_returned_for_correction),
        ("Resolved Corrections", corrections.resolved_corrections),
        ("Unresolved Corrections", corrections.unresolved_corrections),
        ("Correction Completion %", corrections.correction_completion_percentage),
        ("Avg Return to Correction (hrs)", corrections.avg_return_to_correction_hours),
        ("Assessment Correction Rate", corrections.assessment_correction_rate),
    ]:
        ws_c.append([metric, value])
    _auto_width(ws_c, 2)

    # Generate bytes
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    workbook_bytes = buf.getvalue()

    return workbook_bytes


def export_progress_report(
    session: Session,
    ctx: AuthContext | None,
    filter_obj: AnalyticsFilter | None = None,
) -> bytes:
    """Generate an XLSX grading progress report.

    Parameters
    ----------
    session : Session
        Database session.
    ctx : AuthContext | None
        Authorization context.
    filter_obj : AnalyticsFilter | None
        Optional filter.

    Returns
    -------
    bytes
        Workbook bytes.
    """
    from analytics.progress import get_grading_progress

    progress = get_grading_progress(session, ctx, filter_obj=filter_obj)
    metadata = _build_metadata("Grading Progress Report", filter_obj)

    wb = Workbook()
    _write_metadata_sheet(wb, metadata)

    # Progress sheet
    ws = wb.active
    assert ws is not None  # noqa: S101
    ws.title = "Progress"
    headers = ["Metric", "Value"]
    ws.append(headers)
    _style_header(ws, len(headers))
    for metric, value in [
        ("Total Submissions", progress.total_submissions),
        ("Not Started", progress.not_started),
        ("Claimed", progress.claimed),
        ("Draft", progress.draft),
        ("Completed", progress.completed),
        ("Needs Correction", progress.needs_correction),
        ("Corrected", progress.corrected),
        ("Approved", progress.approved),
        ("Finalized", progress.finalized),
        ("% Complete", progress.percentage_complete),
        ("% Approved", progress.percentage_approved),
        ("Active Claims", progress.active_claims),
        ("Stale Claims", progress.stale_claims),
        ("Avg Claim to Completion (hrs)", progress.avg_claim_to_completion_hours),
        ("Avg Return to Correction (hrs)", progress.avg_return_to_correction_hours),
    ]:
        ws.append([metric, value])
    _auto_width(ws, 2)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
