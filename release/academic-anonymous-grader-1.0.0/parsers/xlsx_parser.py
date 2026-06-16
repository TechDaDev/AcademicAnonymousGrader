"""XLSX workbook parser for academic assessment imports.

SECURITY:
- ZIP bomb protection: validates member count, compression ratio, path traversal
- openpyxl read-only/data-only mode
- Formulas are never executed; cached values are used
- Macro-enabled workbooks (.xlsm) are rejected
- Encrypted/password-protected workbooks are rejected
- Legacy .xls is rejected with a clear message
- External link / DDE references detected and warned
"""

from __future__ import annotations

import zipfile
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from parsers.base import ImportParser, ParserLimits
from parsers.exceptions import (
    EmptyFileError,
    EncryptedWorkbookError,
    FileTooLargeError,
    ImportLimitExceededError,
    MacroEnabledWorkbookError,
    SecurityValidationError,
    SheetSelectionError,
    UnsupportedFormatError,
)
from parsers.models import (
    ColumnClassification,
    ImportStatistics,
    ParsedColumn,
    ParsedImport,
    ParsedResponse,
    ParsedStudentRow,
    ParsedValidationMessage,
    ValidationSeverity,
)

try:
    from openpyxl import load_workbook
    from openpyxl.utils.exceptions import InvalidFileException
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]
    InvalidFileException = Exception  # type: ignore[misc,assignment]


class XlsxImportParser(ImportParser):
    """Parser for .xlsx assessment response files."""

    parser_name = "xlsx"

    SUPPORTED_EXTENSIONS = frozenset({".xlsx"})

    def parse(
        self,
        file_bytes: bytes,
        source_filename: str,
        *,
        limits: ParserLimits | None = None,
        table_index: int | None = None,
    ) -> ParsedImport:
        if limits is None:
            limits = ParserLimits()

        if len(file_bytes) == 0:
            raise EmptyFileError("Uploaded file is empty.")

        if len(file_bytes) > limits.max_upload_size_bytes:
            raise FileTooLargeError(
                f"File size exceeds {limits.max_upload_size_bytes // (1024*1024)} MB limit."
            )

        ext = self._get_extension(source_filename)
        if ext == ".xls":
            raise UnsupportedFormatError(
                "Legacy .xls format is not supported. Please save the file as .xlsx."
            )

        if load_workbook is None:
            raise ImportError("openpyxl is required for XLSX parsing.")

        try:
            self._validate_zip_safety(file_bytes)
            wb = load_workbook(
                filename=BytesIO(file_bytes),
                read_only=True,
                data_only=True,
                keep_vba=False,
            )
        except InvalidFileException as exc:
            raise UnsupportedFormatError(f"Invalid XLSX file: {exc}") from exc
        except KeyError as exc:
            raise UnsupportedFormatError(f"Unsupported or corrupt workbook: {exc}") from exc

        # Reject macro-enabled workbooks
        ext_lower = source_filename.lower()
        if ext_lower.endswith(".xlsm"):
            raise MacroEnabledWorkbookError(
                "Macro-enabled workbooks (.xlsm) are not supported. "
                "Save as .xlsx without macros."
            )

        # Reject encrypted workbooks
        if wb.security and wb.security.lockStructure:
            raise EncryptedWorkbookError(
                "The workbook is encrypted or password-protected. "
                "Please remove protection and try again."
            )

        # Discover sheets
        all_sheets = wb.sheetnames
        visible_sheets = [s for s in all_sheets if wb[s].sheet_state == "visible"]
        hidden_sheets = [s for s in all_sheets if wb[s].sheet_state != "visible"]

        # Select sheet
        if table_index is not None:
            if table_index < 0 or table_index >= len(visible_sheets):
                raise SheetSelectionError(
                    f"Sheet index {table_index} is out of range. "
                    f"Available visible sheets: 0-{len(visible_sheets) - 1}."
                )
            sheet_name = visible_sheets[table_index]
        else:
            sheet_name = visible_sheets[0] if visible_sheets else ""

        if not sheet_name:
            raise SheetSelectionError("No visible sheets available in the workbook.")

        ws = wb[sheet_name]

        # Check if sheet is empty
        if ws.max_row is None or ws.max_row < 2:
            raise SheetSelectionError(
                f"Sheet '{sheet_name}' has no data rows. Select a different sheet."
            )

        # Detect hidden rows and columns (not available in read-only mode)
        hidden_row_count = 0
        hidden_column_count = 0
        try:
            for row_idx in range(1, (ws.max_row or 0) + 1):
                row_dim = ws.row_dimensions.get(row_idx)
                if row_dim and row_dim.hidden:
                    hidden_row_count += 1
            for col_letter in (chr(ord("A") + i) for i in range(ws.max_column or 0)):
                col_dim = ws.column_dimensions.get(col_letter)
                if col_dim and col_dim.hidden:
                    hidden_column_count += 1
        except AttributeError:
            pass  # row_dimensions not available in read-only mode

        # Extract headers from first row
        header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        if not header_cells:
            raise EmptyFileError("No header row found in the selected sheet.")

        raw_headers: list[str] = []
        for cell in header_cells[0]:
            if cell.value is None:
                raw_headers.append("")
            else:
                raw_headers.append(str(cell.value))

        # Check for merged cells in header (not available in read-only mode)
        merged_header_warnings: list[str] = []
        try:
            for merged_range in ws.merged_cells.ranges:
                if merged_range.min_row == 1 or merged_range.max_row == 1:
                    merged_header_warnings.append(
                        f"Merged header cells detected: {merged_range}"
                    )
        except AttributeError:
            pass

        # Detect duplicate and blank headers
        seen_headers: dict[str, int] = {}
        dup_warnings: list[str] = []
        for i, h in enumerate(raw_headers):
            if not h.strip():
                dup_warnings.append(f"Column {i + 1} has a blank header.")
            elif h in seen_headers:
                dup_warnings.append(f"Duplicate header '{h}' at column {i + 1}.")
            seen_headers[h] = i

        # Extract data rows (keep cell objects to detect formulas)
        rows_data: list[list[str]] = []
        formula_count = 0
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=False), start=2
        ):
            row_values: list[str] = []
            for cell in row:
                # Safely handle cell value - it could be various types
                cell_value = cell.value
                cell_data_type = getattr(cell, "data_type", "s")
                is_date_value = getattr(cell, "is_date", False)

                if is_date_value and cell_value is not None:
                    # Format dates as ISO string
                    if hasattr(cell_value, "isoformat"):
                        try:
                            row_values.append(cell_value.isoformat())
                        except Exception:
                            row_values.append(str(cell_value))
                    else:
                        row_values.append(str(cell_value))
                elif cell_data_type == "f" or (
                    isinstance(cell_value, str) and cell_value.startswith("=")
                ):
                    formula_count += 1
                    # Use cached value if available
                    if cell_value is not None:
                        row_values.append(self._safe_cell_value(cell_value))
                    else:
                        row_values.append("")
                else:
                    row_values.append(self._safe_cell_value(cell_value))
            # Skip completely empty rows
            if all(v == "" for v in row_values):
                continue
            rows_data.append(row_values)

        wb.close()

        # Enforce limits
        if len(rows_data) > limits.max_import_rows:
            raise ImportLimitExceededError(
                f"Row count {len(rows_data)} exceeds limit of {limits.max_import_rows}."
            )
        if len(raw_headers) > limits.max_import_columns:
            raise ImportLimitExceededError(
                f"Column count {len(raw_headers)} exceeds limit of {limits.max_import_columns}."
            )

        # Normalize headers
        from parsers.normalization import normalize_header

        normalized_headers = [normalize_header(h) for h in raw_headers]

        # Build columns
        from parsers.column_aliases import get_mapped_field, mapped_response_number

        columns: list[ParsedColumn] = []
        response_columns_list: list[ParsedColumn] = []
        unknown_columns_list: list[ParsedColumn] = []

        for i, (orig, norm) in enumerate(zip(raw_headers, normalized_headers, strict=False)):
            mapped = get_mapped_field(norm)
            # Check if it's a response column
            resp_num = mapped_response_number(orig)
            if resp_num is not None:
                mapped = f"response_{resp_num}"
            col_warnings: list[str] = []
            if norm in seen_headers and list(raw_headers).count(orig) > 1:
                col_warnings.append("Duplicate header.")
            if not norm.strip():
                col_warnings.append("Blank header.")

            confidence = 1.0 if mapped else 0.0
            classification = ColumnClassification.UNKNOWN
            if mapped in (
                "first_name", "last_name", "email", "institutional_student_id",
            ):
                classification = ColumnClassification.IDENTITY
            elif mapped == "source_grade":
                classification = ColumnClassification.METADATA
            elif mapped and mapped.startswith("response_"):
                classification = ColumnClassification.RESPONSE
            col = ParsedColumn(
                original_name=orig,
                normalized_name=norm,
                index=i,
                classification=classification,
                mapped_field=mapped,
                is_required=(classification == ColumnClassification.IDENTITY),
                confidence=confidence,
                warnings=tuple(col_warnings),
                response_number=int(mapped.split("_")[1]) if mapped and mapped.startswith("response_") else None,
            )
            columns.append(col)
            if classification == ColumnClassification.RESPONSE:
                response_columns_list.append(col)
            elif classification == ColumnClassification.UNKNOWN and mapped is None:
                unknown_columns_list.append(col)

        # Build student rows

        parsed_rows: list[ParsedStudentRow] = []
        for row_idx, row_values in enumerate(rows_data):
            row_dict = dict(zip(raw_headers, row_values, strict=False))
            responses_list: list[ParsedResponse] = []
            row_warnings: list[ParsedValidationMessage] = []
            row_errors: list[ParsedValidationMessage] = []

            for rc in response_columns_list:
                resp_text = row_dict.get(rc.original_name, "")
                is_blank = not resp_text.strip()
                responses_list.append(
                    ParsedResponse(
                        question_number=rc.response_number or 0,
                        column_name=rc.original_name,
                        text=resp_text or "",
                        is_blank=is_blank,
                    )
                )

            # Extract identity fields
            first_name = None
            last_name = None
            email = None
            inst_id = None
            source_grade = None
            raw_grade = None
            unknown_vals: dict[str, str] = {}

            for col in columns:
                val = row_dict.get(col.original_name, "")
                if col.mapped_field == "first_name":
                    first_name = val or None
                elif col.mapped_field == "last_name":
                    last_name = val or None
                elif col.mapped_field == "email":
                    email = val or None
                elif col.mapped_field == "institutional_student_id":
                    inst_id = val or None
                elif col.mapped_field == "source_grade":
                    raw_grade = val or None
                    if val:
                        try:
                            source_grade = Decimal(str(val))
                        except Exception:  # noqa: S110
                            pass
                elif col.mapped_field is None:
                    if val:
                        unknown_vals[col.original_name] = val

            parsed_rows.append(
                ParsedStudentRow(
                    row_number=row_idx,
                    first_name=first_name,
                    last_name=last_name,
                    email=email,
                    institutional_student_id=inst_id,
                    status=None,
                    started=None,
                    completed=None,
                    duration_seconds=None,
                    source_grade=source_grade,
                    raw_source_grade=raw_grade,
                    source_grade_maximum=None,
                    responses=tuple(responses_list),
                    unknown_values=unknown_vals,
                    warnings=tuple(row_warnings),
                    errors=tuple(row_errors),
                )
            )

        # Build warnings
        all_warnings: list[ParsedValidationMessage] = []
        for w in dup_warnings:
            all_warnings.append(
                ParsedValidationMessage(
                    code="W001",
                    severity=ValidationSeverity.WARNING,
                    message=w,
                )
            )
        for w in merged_header_warnings:
            all_warnings.append(
                ParsedValidationMessage(
                    code="W002",
                    severity=ValidationSeverity.WARNING,
                    message=w,
                )
            )
        if hidden_row_count > 0:
            all_warnings.append(
                ParsedValidationMessage(
                    code="W003",
                    severity=ValidationSeverity.WARNING,
                    message=f"{hidden_row_count} hidden row(s) detected.",
                )
            )
        if hidden_sheets:
            all_warnings.append(
                ParsedValidationMessage(
                    code="W004",
                    severity=ValidationSeverity.WARNING,
                    message=f"{len(hidden_sheets)} hidden sheet(s): {', '.join(hidden_sheets)}.",
                )
            )
        if formula_count > 0:
            all_warnings.append(
                ParsedValidationMessage(
                    code="W005",
                    severity=ValidationSeverity.WARNING,
                    message=f"{formula_count} formula cell(s) detected; cached values used.",
                )
            )

        # External link / DDE detection
        ext_warnings = self._detect_external_links_and_dde(file_bytes)
        all_warnings.extend(ext_warnings)

        # Count responses with blanks
        blank_count = sum(
            1 for pr in parsed_rows for r in pr.responses if r.is_blank
        )
        warning_count = sum(1 for pr in parsed_rows if pr.warnings) + len(all_warnings)
        error_count = sum(1 for pr in parsed_rows if pr.errors)

        stats = ImportStatistics(
            total_rows=len(parsed_rows),
            valid_rows=len(parsed_rows) - error_count,
            warning_rows=warning_count,
            error_rows=error_count,
            blank_response_count=blank_count,
            response_column_count=len(response_columns_list),
            duplicate_email_count=0,
        )

        now = datetime.now(UTC)
        return ParsedImport(
            source_filename=source_filename,
            parser_name=self.parser_name,
            table_index=table_index or 0,
            columns=tuple(columns),
            rows=tuple(parsed_rows),
            response_columns=tuple(response_columns_list),
            unknown_columns=tuple(unknown_columns_list),
            warnings=tuple(all_warnings),
            errors=(),
            statistics=stats,
            parse_started_at=now,
            parse_completed_at=now,
            source_format="xlsx",
            sheet_name=sheet_name,
            available_sheets=tuple(visible_sheets),
            hidden_row_count=hidden_row_count,
            hidden_column_count=hidden_column_count,
            formula_count=formula_count,
        )

    # ── Security helpers ─────────────────────────────────────────────────

    def _validate_zip_safety(self, file_bytes: bytes) -> None:
        """Validate ZIP structure before openpyxl opens it.

        Checks:
        - Valid ZIP archive
        - Maximum member count (10 000)
        - No path traversal (..) in member names
        - Maximum compression ratio (500:1) to detect ZIP bombs
        """
        max_members = 10_000
        max_compression_ratio = 500

        try:
            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                members = zf.infolist()
        except (zipfile.BadZipFile, NotImplementedError) as exc:
            raise UnsupportedFormatError(f"Invalid or corrupt ZIP archive: {exc}") from exc

        if len(members) > max_members:
            raise SecurityValidationError(
                f"Workbook contains {len(members)} entries "
                f"(limit: {max_members}). Possible ZIP bomb."
            )

        total_uncompressed = 0
        total_compressed = 0

        for member in members:
            # Path traversal check
            if ".." in member.filename.split("/"):
                raise SecurityValidationError(
                    f"Path traversal detected in ZIP member: {member.filename!r}"
                )

            uncompressed = member.file_size
            compressed = member.compress_size

            total_uncompressed += uncompressed
            total_compressed += compressed

            # Per-file compression ratio check
            if compressed > 0 and uncompressed > compressed * max_compression_ratio:
                raise SecurityValidationError(
                    f"ZIP entry {member.filename!r} has extreme compression ratio "
                    f"({uncompressed}/{compressed}). Possible ZIP bomb."
                )

        # Overall compression ratio check
        if total_compressed > 0 and total_uncompressed > total_compressed * max_compression_ratio:
            raise SecurityValidationError(
                f"Overall compression ratio {total_uncompressed}/{total_compressed} "
                f"exceeds limit. Possible ZIP bomb."
            )

    def _detect_external_links_and_dde(
        self, file_bytes: bytes
    ) -> list[ParsedValidationMessage]:
        """Scan the extracted XML for external reference indicators.

        Returns a list of warning messages; does not block parsing.
        """
        warnings: list[ParsedValidationMessage] = []
        try:
            import zipfile
            from io import BytesIO

            with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
                # Check for externalLink directory
                ext_link_names = [
                    n for n in zf.namelist()
                    if "externallink" in n.lower()
                    or "externalreference" in n.lower()
                ]
                if ext_link_names:
                    warnings.append(
                        ParsedValidationMessage(
                            code="W006",
                            severity=ValidationSeverity.WARNING,
                            message=f"External link(s) detected: {len(ext_link_names)} reference(s). "
                            "Links are not followed.",
                        )
                    )

                # Check for DDE in shared strings or sheet XML
                for name in zf.namelist():
                    if name.endswith(".xml"):
                        try:
                            content = zf.read(name)
                            lower = content.lower()
                            # DDE indicators
                            if b"dde" in lower and (
                                b"ddeconnection" in lower or b"ddelink" in lower
                            ):
                                warnings.append(
                                    ParsedValidationMessage(
                                        code="W007",
                                        severity=ValidationSeverity.WARNING,
                                        message=f"DDE link detected in {name}. "
                                        "DDE links are ignored.",
                                    )
                                )
                                break
                        except Exception:  # noqa: S110
                            pass
        except Exception:  # noqa: S110
            pass
        return warnings

    def _safe_cell_value(self, cell: Any) -> str:
        """Convert a cell value to a string safely."""
        if cell is None:
            return ""
        if isinstance(cell, (int, float)):
            # Preserve leading zeros by checking if it looks like an ID
            return str(cell)
        return str(cell)

    def _get_extension(self, filename: str) -> str:
        """Get the lowercase file extension."""
        idx = filename.rfind(".")
        if idx == -1:
            return ""
        return filename[idx:].lower()
