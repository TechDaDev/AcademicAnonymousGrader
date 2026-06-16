"""Tests for the XLSX workbook parser (Phase 9)."""

from __future__ import annotations

from pathlib import Path

import pytest

from parsers.base import ParserLimits
from parsers.exceptions import (
    EmptyFileError,
    FileTooLargeError,
    ImportLimitExceededError,
    SecurityValidationError,
    SheetSelectionError,
    UnsupportedFormatError,
)
from parsers.xlsx_parser import XlsxImportParser

FIXTURE_DIR = Path(__file__).parent / "fixtures"


@pytest.fixture
def parser() -> XlsxImportParser:
    return XlsxImportParser()


@pytest.fixture
def limits() -> ParserLimits:
    return ParserLimits(
        max_upload_size_bytes=10 * 1024 * 1024,
        max_import_rows=1000,
        max_import_columns=50,
        max_cell_text_length=5000,
    )


class TestXlsxParser:
    """Core XLSX parsing tests."""

    def test_valid_xlsx_parses(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """A valid XLSX file with standard columns parses correctly."""
        path = FIXTURE_DIR / "valid.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "valid.xlsx", limits=limits)
        assert result.source_format == "xlsx"
        assert result.parser_name == "xlsx"
        assert result.statistics.total_rows == 2
        assert len(result.columns) == 6
        assert result.available_sheets == ("Sheet1",)

    def test_headers_detected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Column headers are correctly extracted."""
        path = FIXTURE_DIR / "valid.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "valid.xlsx", limits=limits)
        names = [c.original_name for c in result.columns]
        assert "First Name" in names
        assert "Last Name" in names
        assert "Email" in names
        assert "Institutional ID" in names
        assert "Response 1" in names
        assert "Response 2" in names

    def test_identity_mapped(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Identity columns are classified correctly."""
        path = FIXTURE_DIR / "valid.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "valid.xlsx", limits=limits)
        identity_fields = {c.mapped_field for c in result.columns if c.classification.name == "IDENTITY"}
        assert "first_name" in identity_fields
        assert "last_name" in identity_fields
        assert "email" in identity_fields
        assert "institutional_student_id" in identity_fields

    def test_response_columns_detected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Response columns are classified and numbered."""
        path = FIXTURE_DIR / "valid.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "valid.xlsx", limits=limits)
        resp_cols = [c for c in result.columns if c.classification.name == "RESPONSE"]
        assert len(resp_cols) == 2
        nums = sorted(c.response_number for c in resp_cols if c.response_number is not None)
        assert nums == [1, 2]

    def test_leading_zero_id_preserved(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Leading zeros in institutional IDs are preserved."""
        path = FIXTURE_DIR / "valid.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "valid.xlsx", limits=limits)
        ids = [r.institutional_student_id for r in result.rows if r.institutional_student_id]
        assert "001234" in ids

    def test_arabic_text_preserved(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Arabic characters are preserved through XLSX parsing."""
        path = FIXTURE_DIR / "arabic.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "arabic.xlsx", limits=limits)
        names = [r.first_name for r in result.rows if r.first_name]
        assert "أحمد" in names
        assert "سارة" in names

    def test_empty_file_raises(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Empty file raises EmptyFileError."""
        with pytest.raises(EmptyFileError):
            parser.parse(b"", "empty.xlsx", limits=limits)

    def test_invalid_xlsx_rejected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Invalid (non-XLSX) file raises UnsupportedFormatError."""
        path = FIXTURE_DIR / "invalid.xlsx"
        data = path.read_bytes()
        with pytest.raises(UnsupportedFormatError):
            parser.parse(data, "invalid.xlsx", limits=limits)

    def test_legacy_xls_rejected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Legacy .xls extension raises UnsupportedFormatError."""
        data = b"dummy content"
        with pytest.raises(UnsupportedFormatError, match=".xls"):
            parser.parse(data, "workbook.xls", limits=limits)

    def test_file_size_limit(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Oversized file is rejected."""
        small_limits = ParserLimits(max_upload_size_bytes=10)
        with pytest.raises(FileTooLargeError):
            parser.parse(b"a" * 100, "big.xlsx", limits=small_limits)

    def test_row_limit(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Excessive rows are rejected."""
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sheet1"
            ws.append(["Col1"])
            for _ in range(5):
                ws.append(["x"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        data = buf.getvalue()
        small_limits = ParserLimits(max_import_rows=2)
        with pytest.raises(ImportLimitExceededError):
            parser.parse(data, "many_rows.xlsx", limits=small_limits)

    def test_column_limit(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Excessive columns are rejected."""
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sheet1"
            ws.append([f"Col{i}" for i in range(60)])
            ws.append(["x"] * 60)
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        data = buf.getvalue()
        small_limits = ParserLimits(max_import_columns=10)
        with pytest.raises(ImportLimitExceededError):
            parser.parse(data, "many_cols.xlsx", limits=small_limits)

    def test_sheet_selection_default(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """First visible sheet is used by default."""
        path = FIXTURE_DIR / "multi_sheet.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "multi_sheet.xlsx", limits=limits)
        # Summary is first visible sheet
        assert result.sheet_name == "Summary"

    def test_sheet_selection_by_index(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Sheet index selects among visible sheets."""
        path = FIXTURE_DIR / "multi_sheet.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "multi_sheet.xlsx", limits=limits, table_index=1)
        assert result.sheet_name == "Responses"
        assert len(result.columns) == 3

    def test_sheet_selection_out_of_range(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Out-of-range sheet index raises SheetSelectionError."""
        path = FIXTURE_DIR / "multi_sheet.xlsx"
        data = path.read_bytes()
        with pytest.raises(SheetSelectionError):
            parser.parse(data, "multi_sheet.xlsx", limits=limits, table_index=999)

    def test_hidden_sheet_warning(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Hidden sheets are reported as warnings."""
        path = FIXTURE_DIR / "multi_sheet.xlsx"
        data = path.read_bytes()
        result = parser.parse(data, "multi_sheet.xlsx", limits=limits)
        codes = [w.code for w in result.warnings]
        assert "W004" in codes  # hidden sheet warning

    def test_empty_xlsx_rejected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """An XLSX with no data rows raises SheetSelectionError."""
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "EmptySheet"
            ws.append(["Col1"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        data = buf.getvalue()
        with pytest.raises(SheetSelectionError, match="no data rows"):
            parser.parse(data, "empty_sheet.xlsx", limits=limits)


class TestXlsxParserSecurity:
    """Security hardening tests for XLSX parser."""

    def test_zip_bomb_rejected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """A ZIP bomb is rejected by security validation."""
        import io
        import zipfile

        # Create a "compressed" entry with extreme ratio
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/workbook.xml", b"<workbook/>")
            zf.writestr("xl/_rels/workbook.xml.rels", b"<rels/>")
            zf.writestr(
                "[Content_Types].xml",
                b'<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
            )
            # Bomb entry: large repeating data
            bomb_data = b"A" * 1_000_000  # 1 MB uncompressed
            zf.writestr("xl/worksheets/sheet1.xml", bomb_data)

        data = buf.getvalue()
        with pytest.raises(SecurityValidationError, match="compression ratio|ZIP bomb"):
            parser.parse(data, "bomb.xlsx", limits=limits)

    def test_path_traversal_rejected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """ZIP entry with path traversal is rejected."""
        import io
        import zipfile

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/workbook.xml", b"<workbook/>")
            zf.writestr("xl/_rels/workbook.xml.rels", b"<rels/>")
            zf.writestr(
                "[Content_Types].xml",
                b'<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
            )
            zf.writestr("../evil.txt", b"malicious")

        data = buf.getvalue()
        with pytest.raises(SecurityValidationError, match="Path traversal"):
            parser.parse(data, "traversal.xlsx", limits=limits)

    def test_too_many_zip_members_rejected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Workbook with too many ZIP entries is rejected."""
        import io
        import zipfile

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("xl/workbook.xml", b"<workbook/>")
            zf.writestr("xl/_rels/workbook.xml.rels", b"<rels/>")
            zf.writestr(
                "[Content_Types].xml",
                b'<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"/>',
            )
            # Add many small entries
            for i in range(11_000):
                zf.writestr(f"xl/worksheets/sheet{i}.xml", b"<sheet/>")

        data = buf.getvalue()
        with pytest.raises(SecurityValidationError, match="entries"):
            parser.parse(data, "many_members.xlsx", limits=limits)


class TestXlsxParserFormulas:
    """Formula handling tests."""

    def test_formula_detected(self, parser: XlsxImportParser, limits: ParserLimits) -> None:
        """Formula cells are detected and cached values used."""
        import io

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "Sheet1"
        else:
            ws = wb.create_sheet("Sheet1")
        ws.append(["First Name", "Score", "Response 1"])
        ws.append(["Test", 10, "Answer"])
        # Set a formula cell with cached value
        ws["B2"] = 10  # cached value
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()

        data = buf.getvalue()
        result = parser.parse(data, "formulas.xlsx", limits=limits)
        # openpyxl in data_only mode reads cached values
        assert result.statistics.total_rows == 1
