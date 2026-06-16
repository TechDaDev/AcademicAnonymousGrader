"""Excel export security and privacy audit tests."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import openpyxl
from sqlalchemy.orm import Session

from models.anonymous_student import AnonymousStudent
from models.assessment import Assessment
from models.grade_record import GradeRecord
from models.import_batch import ImportBatch
from models.material import Material
from models.question import Question
from models.student_identity import StudentIdentity
from models.submission import Submission
from security.encryption import encrypt_text
from security.key_validation import _decode_key
from security.models import EncryptionKey
from services.authorization_service import AuthContext
from services.excel_export_service import generate_export_workbook
from services.finalization_service import finalize_assessment

TEST_EKEY = "8k7lYYckaMaqKDy31LgQhPCTvSup2elJtoEm0TEztXY="
TEST_FKEY = "yQH1YtA5kQtrilNmVo_qxLT0Yty4sl4k5vZMDnxgp-c="


def _setup(session: Session, settings) -> str:
    key = EncryptionKey(_decode_key(settings.identity_encryption_key, "IDENTITY_ENCRYPTION_KEY", 32))
    material = Material(name="Sec Test"); session.add(material); session.flush()
    assessment = Assessment(material_id=material.id, title="Sec", maximum_grade=Decimal("100"))
    session.add(assessment); session.flush()
    q1 = Question(assessment_id=assessment.id, question_number=1, maximum_grade=Decimal("100"))
    session.add(q1); session.flush()
    batch = ImportBatch(assessment_id=assessment.id, source_filename="t.html"); session.add(batch); session.flush()
    identity = StudentIdentity(
        encrypted_first_name=encrypt_text(key, "Security"),
        encrypted_last_name=encrypt_text(key, "Test"),
        encrypted_email=encrypt_text(key, "sec@example.com"),
    )
    session.add(identity); session.flush()
    anon = AnonymousStudent(student_identity_id=identity.id, anonymous_code="STU-SEC01")
    session.add(anon); session.flush()
    sub = Submission(assessment_id=assessment.id, anonymous_student_id=anon.id, import_batch_id=batch.id, review_status="approved")
    session.add(sub); session.flush()
    gr = GradeRecord(submission_id=sub.id, question_id=q1.id, grade=Decimal("85"), grading_status="graded", feedback="Good work")
    session.add(gr); session.flush()
    finalize_assessment(session, assessment.id, auth_ctx=_ADMIN_AUTH)
    return assessment.id


class TestExportSecurity:
    def test_no_ciphertext_exported(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        aid = _setup(session, settings)
        result = generate_export_workbook(session, aid, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and "enc:" in cell:
                        assert False, f"Ciphertext found in {sheet_name}: {cell[:50]}"

    def test_no_encryption_key_exported(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        aid = _setup(session, settings)
        result = generate_export_workbook(session, aid, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and TEST_EKEY[:10] in cell:
                        assert False, "Encryption key fragment found in workbook"

    def test_source_grade_not_used(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        key = EncryptionKey(_decode_key(settings.identity_encryption_key, "IDENTITY_ENCRYPTION_KEY", 32))
        material = Material(name="Src Test"); session.add(material); session.flush()
        assessment = Assessment(material_id=material.id, title="Src", maximum_grade=Decimal("100"))
        session.add(assessment); session.flush()
        q1 = Question(assessment_id=assessment.id, question_number=1, maximum_grade=Decimal("100"))
        session.add(q1); session.flush()
        batch = ImportBatch(assessment_id=assessment.id, source_filename="t.html"); session.add(batch); session.flush()
        identity = StudentIdentity(encrypted_first_name=encrypt_text(key, "Src"))
        session.add(identity); session.flush()
        anon = AnonymousStudent(student_identity_id=identity.id, anonymous_code="STU-SRC01")
        session.add(anon); session.flush()
        sub = Submission(assessment_id=assessment.id, anonymous_student_id=anon.id, import_batch_id=batch.id, review_status="approved", source_grade=Decimal("999"))
        session.add(sub); session.flush()
        gr = GradeRecord(submission_id=sub.id, question_id=q1.id, grade=Decimal("85"), grading_status="graded")
        session.add(gr); session.flush()
        finalize_assessment(session, assessment.id, auth_ctx=_ADMIN_AUTH)
        result = generate_export_workbook(session, assessment.id, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        grade = wb["Final Grades"].cell(row=2, column=7).value
        assert grade == 85.0  # GradeRecord value, not source_grade

    def test_formula_injection_prevented(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        key = EncryptionKey(_decode_key(settings.identity_encryption_key, "IDENTITY_ENCRYPTION_KEY", 32))
        material = Material(name="Formula Test"); session.add(material); session.flush()
        assessment = Assessment(material_id=material.id, title="Formula", maximum_grade=Decimal("100"))
        session.add(assessment); session.flush()
        q1 = Question(assessment_id=assessment.id, question_number=1, maximum_grade=Decimal("100"))
        session.add(q1); session.flush()
        batch = ImportBatch(assessment_id=assessment.id, source_filename="t.html"); session.add(batch); session.flush()
        identity = StudentIdentity(
            encrypted_first_name=encrypt_text(key, "=SUM(A1:A10)"),
            encrypted_last_name=encrypt_text(key, "Test"),
            encrypted_email=encrypt_text(key, "formula@example.com"),
        )
        session.add(identity); session.flush()
        anon = AnonymousStudent(student_identity_id=identity.id, anonymous_code="STU-FORM01")
        session.add(anon); session.flush()
        sub = Submission(assessment_id=assessment.id, anonymous_student_id=anon.id, import_batch_id=batch.id, review_status="approved")
        session.add(sub); session.flush()
        gr = GradeRecord(submission_id=sub.id, question_id=q1.id, grade=Decimal("85"), grading_status="graded")
        session.add(gr); session.flush()
        finalize_assessment(session, assessment.id, auth_ctx=_ADMIN_AUTH)
        result = generate_export_workbook(session, assessment.id, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        first_name_cell = wb["Final Grades"].cell(row=2, column=2).value
        assert first_name_cell is not None
        assert str(first_name_cell).startswith("'"), f"Formula not protected: {first_name_cell}"

    def test_identities_only_in_intended_columns(self, session: Session, monkeypatch) -> None:
        """Verify decrypted identities appear only in identity columns."""
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        key = EncryptionKey(_decode_key(settings.identity_encryption_key, "IDENTITY_ENCRYPTION_KEY", 32))
        material = Material(name="Audit"); session.add(material); session.flush()
        assessment = Assessment(material_id=material.id, title="Audit", maximum_grade=Decimal("100"))
        session.add(assessment); session.flush()
        q1 = Question(assessment_id=assessment.id, question_number=1, maximum_grade=Decimal("100"))
        session.add(q1); session.flush()
        batch = ImportBatch(assessment_id=assessment.id, source_filename="t.html"); session.add(batch); session.flush()
        identity = StudentIdentity(
            encrypted_first_name=encrypt_text(key, "FirstName_Only"),
            encrypted_last_name=encrypt_text(key, "LastName_Only"),
            encrypted_email=encrypt_text(key, "audit@example.com"),
        )
        session.add(identity); session.flush()
        anon = AnonymousStudent(student_identity_id=identity.id, anonymous_code="STU-AUDIT01")
        session.add(anon); session.flush()
        sub = Submission(assessment_id=assessment.id, anonymous_student_id=anon.id, import_batch_id=batch.id, review_status="approved")
        session.add(sub); session.flush()
        gr = GradeRecord(submission_id=sub.id, question_id=q1.id, grade=Decimal("85"), grading_status="graded")
        session.add(gr); session.flush()
        finalize_assessment(session, assessment.id, auth_ctx=_ADMIN_AUTH)
        result = generate_export_workbook(session, assessment.id, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        ws = wb["Final Grades"]
        # Identity values in identity columns (2,3,4,5) are fine; check non-identity columns
        anon_code_col6 = ws.cell(row=2, column=6).value
        assert anon_code_col6 == "STU-AUDIT01"
        # Check no identity values appear in non-identity columns (7+)
        for col in range(7, 15):
            val = ws.cell(row=2, column=col).value
            if val and isinstance(val, str):
                assert "FirstName_Only" not in val, f"Identity leaked in col {col}"
                assert "LastName_Only" not in val, f"Identity leaked in col {col}"

    def test_no_hidden_sheets(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        aid = _setup(session, settings)
        result = generate_export_workbook(session, aid, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            assert ws.sheet_state == "visible", f"Hidden sheet: {sheet_name}"

    def test_workbook_metadata_no_secrets(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        aid = _setup(session, settings)
        result = generate_export_workbook(session, aid, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        props = wb.properties
        assert props.creator == "Academic Anonymous Grader"
        # No identity data in metadata
        for attr in ("creator", "title", "description", "subject", "keywords"):
            val = getattr(props, attr, None)
            if val and isinstance(val, str):
                assert "Student" not in val or "Grader" in val
                assert "Identity" not in val or "Encryption" not in val

    def test_no_student_id_uuid_in_workbook(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        aid = _setup(session, settings)
        result = generate_export_workbook(session, aid, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str):
                        # UUIDs are 36 chars with hyphens
                        if len(cell) == 36 and cell.count("-") == 4:
                            assert False, f"UUID found in {sheet_name}: {cell}"


_ADMIN_AUTH = AuthContext(user_id="test-admin", role="administrator")
