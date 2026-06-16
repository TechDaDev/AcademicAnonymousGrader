"""Excel export service tests."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from sqlalchemy.orm import Session

from models.anonymous_student import AnonymousStudent
from models.assessment import Assessment
from models.grade_record import GradeRecord
from models.import_batch import ImportBatch
from models.material import Material
from models.question import Question
from models.student_identity import StudentIdentity
from models.submission import Submission
from security.encryption import encrypt_text
from security.key_validation import _decode_key
from security.models import EncryptionKey
from services.authorization_service import AuthContext
from services.excel_export_service import generate_export_workbook
from services.exceptions import FinalizedAssessmentExportError
from services.finalization_service import finalize_assessment

TEST_EKEY = "8k7lYYckaMaqKDy31LgQhPCTvSup2elJtoEm0TEztXY="
TEST_FKEY = "yQH1YtA5kQtrilNmVo_qxLT0Yty4sl4k5vZMDnxgp-c="


def _setup(session: Session, settings) -> dict:  # type: ignore[type-arg]
    key = EncryptionKey(_decode_key(settings.identity_encryption_key, "IDENTITY_ENCRYPTION_KEY", 32))
    material = Material(name="XL Test")
    session.add(material); session.flush()
    assessment = Assessment(material_id=material.id, title="XL", maximum_grade=Decimal("100"))
    session.add(assessment); session.flush()
    q1 = Question(assessment_id=assessment.id, question_number=1, maximum_grade=Decimal("40"))
    q2 = Question(assessment_id=assessment.id, question_number=2, maximum_grade=Decimal("60"))
    session.add_all([q1, q2]); session.flush()
    batch = ImportBatch(assessment_id=assessment.id, source_filename="t.html")
    session.add(batch); session.flush()
    identity = StudentIdentity(
        encrypted_first_name=encrypt_text(key, "Ali"),
        encrypted_last_name=encrypt_text(key, "Omar"),
        encrypted_email=encrypt_text(key, "ali@example.com"),
    )
    session.add(identity); session.flush()
    anon = AnonymousStudent(student_identity_id=identity.id, anonymous_code="STU-XL001")
    session.add(anon); session.flush()
    sub = Submission(
        assessment_id=assessment.id, anonymous_student_id=anon.id,
        import_batch_id=batch.id, review_status="approved",
    )
    session.add(sub); session.flush()
    g1 = GradeRecord(submission_id=sub.id, question_id=q1.id, grade=Decimal("35"), grading_status="graded")
    g2 = GradeRecord(submission_id=sub.id, question_id=q2.id, grade=Decimal("55"), grading_status="graded")
    session.add_all([g1, g2]); session.flush()
    finalize_assessment(session, assessment.id, auth_ctx=_ADMIN_AUTH)
    return {"assessment": assessment}


class TestExcelExport:
    def test_workbook_generated(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        data = _setup(session, settings)
        result = generate_export_workbook(session, data["assessment"].id, settings, auth_ctx=_ADMIN_AUTH)
        assert result.workbook_bytes is not None
        assert result.file_hash is not None
        assert result.file_size > 0
        assert result.export_reference.startswith("EXP-")

    def test_workbook_opens(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        data = _setup(session, settings)
        result = generate_export_workbook(session, data["assessment"].id, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        assert wb is not None

    def test_required_sheets(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        data = _setup(session, settings)
        result = generate_export_workbook(session, data["assessment"].id, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        for sheet in ("Final Grades", "Question Grades", "Feedback", "Export Summary"):
            assert sheet in wb.sheetnames

    def test_row_count(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        data = _setup(session, settings)
        result = generate_export_workbook(session, data["assessment"].id, settings, auth_ctx=_ADMIN_AUTH)
        assert result.row_count == 1

    def test_not_finalized_raises(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        data = _setup(session, settings)
        a = session.query(Assessment).filter(Assessment.id == data["assessment"].id).first()
        assert a is not None
        a.finalization_status = "not_ready"; session.flush()
        with pytest.raises(FinalizedAssessmentExportError):
            generate_export_workbook(session, data["assessment"].id, settings, auth_ctx=_ADMIN_AUTH)

    def test_grades_numeric(self, session: Session, monkeypatch) -> None:
        monkeypatch.setenv("IDENTITY_ENCRYPTION_KEY", TEST_EKEY)
        monkeypatch.setenv("IDENTITY_FINGERPRINT_KEY", TEST_FKEY)
        from config import get_settings
        settings = get_settings()
        data = _setup(session, settings)
        result = generate_export_workbook(session, data["assessment"].id, settings, auth_ctx=_ADMIN_AUTH)
        wb = openpyxl.load_workbook(BytesIO(result.workbook_bytes))
        grade = wb["Final Grades"].cell(row=2, column=7).value
        assert isinstance(grade, (int, float))

_ADMIN_AUTH = AuthContext(user_id="test-admin", role="administrator")
