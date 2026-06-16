"""Generate synthetic Phase 9 test fixtures.

Run once before Phase 9 tests to create fixture files.
"""
import csv
import io
from pathlib import Path

FIXTURE_DIR = Path(__file__).parent


def _xlsx_bytes(rows: list[list[str]], sheet_name: str = "Sheet1") -> bytes:
    """Build a simple XLSX workbook in memory and return bytes."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(title=sheet_name)
    for row in rows:
        ws.append([str(v) if v is not None else "" for v in row])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


# ── CSV fixtures ──────────────────────────────────────────────────────────

def _write_csv(filename: str, rows: list[list[str]], delimiter: str = ",") -> Path:
    path = FIXTURE_DIR / filename
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=delimiter)
        for row in rows:
            writer.writerow(row)
    return path


def _write_xlsx(filename: str, rows: list[list[str]], sheet_name: str = "Sheet1") -> Path:
    """Write a synthetic XLSX file."""
    path = FIXTURE_DIR / filename
    data = _xlsx_bytes(rows, sheet_name)
    path.write_bytes(data)
    return path


# Valid UTF-8 CSV
DATA_VALID = [
    ["First Name", "Last Name", "Email", "Institutional ID", "Response 1", "Response 2"],
    ["Ali", "Omar", "ali@test.com", "001234", "Answer 1", "Answer 2"],
    ["Sara", "Test", "sara@test.com", "", "Answer 3", "Answer 4"],
]
_write_csv("valid_utf8.csv", DATA_VALID)

# Valid UTF-8 BOM CSV — write with BOM
bom_path = FIXTURE_DIR / "valid_utf8_bom.csv"
with open(bom_path, "w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f)
    for row in DATA_VALID:
        writer.writerow(row)

# Semicolon-delimited CSV
_write_csv("valid_semicolon.csv", DATA_VALID, delimiter=";")

# Tab-delimited CSV
_write_csv("valid_tab.csv", DATA_VALID, delimiter="\t")

# Pipe-delimited CSV
_write_csv("valid_pipe.csv", DATA_VALID, delimiter="|")

# CSV with Arabic text
DATA_ARABIC = [
    ["First Name", "Last Name", "Email", "Response 1"],
    ["أحمد", "محمد", "ahmed@test.com", "إجابة 1"],
    ["سارة", "علي", "sara@test.com", "إجابة 2"],
]
_write_csv("arabic_utf8.csv", DATA_ARABIC)

# CSV with leading-zero ID
DATA_LEADING_ZERO = [
    ["First Name", "Last Name", "Email", "Institutional ID", "Response 1"],
    ["Test", "User", "test@test.com", "001234", "Some response"],
]
_write_csv("leading_zero_id.csv", DATA_LEADING_ZERO)

# CSV with quoted multiline response
DATA_MULTILINE = [
    ["First Name", "Last Name", "Email", "Response 1"],
    ["Multi", "Line", "multi@test.com", "Line 1\nLine 2\nLine 3"],
]
_write_csv("multiline_quoted.csv", DATA_MULTILINE)

# CSV with duplicate headers
DATA_DUP_HEADER = [
    ["First Name", "First Name", "Email", "Response 1"],
    ["A", "B", "a@b.com", "Answer"],
]
_write_csv("duplicate_headers.csv", DATA_DUP_HEADER)

# CSV with blank header
DATA_BLANK_HEADER = [
    ["First Name", "", "Email", "Response 1"],
    ["A", "X", "a@b.com", "Answer"],
]
_write_csv("blank_header.csv", DATA_BLANK_HEADER)

# Malformed CSV (binary content)
with open(FIXTURE_DIR / "malformed.csv", "wb") as bf:
    bf.write(b"\x00\x01\x02\x03\x04This is not a proper CSV\xff\xfe")

# Minimal 1-row CSV (headers only, no data)
_write_csv("headers_only.csv", [["First Name", "Email", "Response 1"]])

# ── XLSX fixtures ────────────────────────────────────────────────────────

# Valid XLSX
DATA_XLSX_VALID = [
    ["First Name", "Last Name", "Email", "Institutional ID", "Response 1", "Response 2"],
    ["Ahmed", "Ali", "ahmed@test.com", "001234", "Answer A1", "Answer A2"],
    ["Sara", "Omar", "sara@test.com", "S567", "Answer B1", "Answer B2"],
]
_write_xlsx("valid.xlsx", DATA_XLSX_VALID)

# Arabic XLSX
DATA_ARABIC_XLSX = [
    ["First Name", "Last Name", "Email", "Response 1"],
    ["أحمد", "محمد", "ahmed@test.com", "إجابة 1"],
    ["سارة", "علي", "sara@test.com", "إجابة 2"],
]
_write_xlsx("arabic.xlsx", DATA_ARABIC_XLSX)

# Multi-sheet XLSX
def _write_multi_sheet_xlsx() -> Path:
    from openpyxl import Workbook

    path = FIXTURE_DIR / "multi_sheet.xlsx"
    wb = Workbook()
    ws1 = wb.active
    if ws1 is not None:
        ws1.title = "Summary"
    else:
        ws1 = wb.create_sheet("Summary")
    ws1.append(["Header1", "Header2"])
    ws1.append(["Data1", "Data2"])

    ws2 = wb.create_sheet("Responses")
    ws2.append(["First Name", "Last Name", "Response 1"])
    ws2.append(["Test", "User", "Answer"])

    ws3 = wb.create_sheet("Hidden")
    ws3.sheet_state = "hidden"
    ws3.append(["Hidden", "Data"])

    wb.save(path)
    wb.close()
    return path

_write_multi_sheet_xlsx()

# XLSX with formulas (cached values only)
def _write_formula_xlsx() -> Path:
    from openpyxl import Workbook

    path = FIXTURE_DIR / "formula.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Sheet1"
    else:
        ws = wb.create_sheet("Sheet1")
    ws.append(["First Name", "Score", "Response 1"])
    ws.append(["Test", 10, "Answer"])
    ws["B2"] = 10  # cached value; formula not stored
    wb.save(path)
    wb.close()
    return path

_write_formula_xlsx()

# Invalid XLSX (not a valid ZIP)
(FIXTURE_DIR / "invalid.xlsx").write_bytes(b"This is not a valid XLSX file")

print(f"Generated fixtures in {FIXTURE_DIR}")
for fp in sorted(FIXTURE_DIR.glob("*")):
    if fp.suffix in (".csv", ".xlsx", ".py"):
        print(f"  {fp.name} ({fp.stat().st_size} B)")


print(f"Generated {len(list(FIXTURE_DIR.glob('*.csv')))} CSV fixtures in {FIXTURE_DIR}")
