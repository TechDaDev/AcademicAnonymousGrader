# Academic Anonymous Grader — Excel Export Service
"""Generate Excel workbooks for finalized assessments.

SECURITY:
- Identity values are decrypted ONLY during workbook generation.
- Never write decrypted values to the database.
- Formula injection is prevented by prefixing dangerous cells.
- No ciphertext, fingerprints, or encryption keys are exported.
"""

from __future__ import annotations

import hashlib
import uuid
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import TYPE_CHECKING, Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.assessment import Assessment
from models.grade_record import GradeRecord
from models.question import Question
from models.submission import Submission
from services.exceptions import (
    ExportValidationError,
    ExportWorkbookError,
    FinalizedAssessmentExportError,
)
from services.export_identity_service import get_export_identity
from services.logging_service import get_logger

if TYPE_CHECKING:
    from sqlalchemy.orm import Session

    from config import Settings

logger = get_logger("excel_export")


# ── Constants ─────────────────────────────────────────────────────

SCHEMA_VERSION = "1.0"
_SAFE_PREFIX = "'"  # Excel formula injection protection prefix

# Header styling
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
_HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=False)


@dataclass(frozen=True, slots=True)
class ExportResult:
    """Result of a workbook generation."""

    workbook_bytes: bytes
    file_hash: str
    file_size: int
    row_count: int
    export_reference: str


# ── Formula injection protection ──────────────────────────────────


def _safe_value(value: Any) -> Any:  # noqa: ANN401
    """Prevent Excel formula injection by prefixing dangerous values."""
    if isinstance(value, str) and value and value[0] in ("=", "+", "-", "@"):
        return _SAFE_PREFIX + value
    return value


# ── Workbook generation ───────────────────────────────────────────


def _get_questions(session: Session, assessment_id: str) -> list[Question]:
    return (
        session.query(Question)
        .filter(Question.assessment_id == assessment_id)
        .order_by(Question.question_number)
        .all()
    )


def _get_submissions(session: Session, assessment_id: str) -> list[Submission]:
    return (
        session.query(Submission)
        .filter(Submission.assessment_id == assessment_id)
        .options(
            __import__("sqlalchemy").orm.joinedload(Submission.anonymous_student),
        )
        .order_by(Submission.id)
        .all()
    )


def _get_grade_map(session: Session, submission_id: str) -> dict[str, GradeRecord]:
    records = (
        session.query(GradeRecord)
        .filter(GradeRecord.submission_id == submission_id)
        .all()
    )
    return {r.question_id: r for r in records}


def _style_header(ws: Any, col_count: int) -> None:  # noqa: ANN401
    """Apply header styling and freeze panes."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGNMENT
    ws.freeze_panes = "A2"


def _auto_width(ws: Any, col_count: int, max_width: int = 40) -> None:  # noqa: ANN401
    """Set reasonable column widths."""
    for col in range(1, col_count + 1):
        max_len = 0
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max(max_len + 2, 10)


# ── Sheet builders ────────────────────────────────────────────────


def _build_final_grades_sheet(
    ws: Any,  # noqa: ANN401
    session: Session,
    assessment: Assessment,
    settings: Settings,
    questions: list[Question],
    submissions: list[Submission],
) -> int:
    """Build the Final Grades sheet."""
    headers = [
        "Institutional Student ID",
        "First Name",
        "Last Name",
        "Full Name",
        "Email",
        "Anonymous Code",
        "Final Grade",
        "Maximum Grade",
        "Percentage",
        "Review Status",
        "Assessment",
        "Material",
        "Academic Year",
        "Exported At",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    now_str = datetime.now(UTC).isoformat()
    material_name = assessment.material.name if assessment.material else ""

    row_num = 1
    for sub in submissions:
        row_num += 1
        anon = sub.anonymous_student
        if anon is None:
            continue

        # Decrypt identity
        try:
            ei = get_export_identity(session, anon.id, settings)
        except Exception:
            ei = None

        grade_map = _get_grade_map(session, sub.id)
        eg_vals: list[Decimal] = [g.grade for g in grade_map.values() if g.grade is not None]
        total_g = sum(eg_vals, Decimal("0"))

        pct = float(total_g) / float(assessment.maximum_grade) if assessment.maximum_grade > 0 else 0.0

        ws.cell(row=row_num, column=1, value=_safe_value(ei.institutional_student_id if ei else None))
        ws.cell(row=row_num, column=2, value=_safe_value(ei.first_name if ei else ""))
        ws.cell(row=row_num, column=3, value=_safe_value(ei.last_name if ei else ""))
        ws.cell(row=row_num, column=4, value=_safe_value(ei.full_name if ei else ""))
        ws.cell(row=row_num, column=5, value=_safe_value(ei.email if ei else ""))
        ws.cell(row=row_num, column=6, value=anon.anonymous_code)
        ws.cell(row=row_num, column=7, value=float(total_g))
        ws.cell(row=row_num, column=8, value=float(assessment.maximum_grade))

        pct_cell = ws.cell(row=row_num, column=9, value=pct)
        pct_cell.number_format = "0.00%"

        ws.cell(row=row_num, column=10, value=sub.review_status)
        ws.cell(row=row_num, column=11, value=_safe_value(assessment.title))
        ws.cell(row=row_num, column=12, value=_safe_value(material_name))
        ws.cell(row=row_num, column=13, value=_safe_value(assessment.academic_year or ""))
        ws.cell(row=row_num, column=14, value=now_str)

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num}"
    _auto_width(ws, len(headers))
    return row_num - 1  # row count (excl header)


def _build_question_grades_sheet(
    ws: Any,  # noqa: ANN401
    session: Session,
    assessment: Assessment,
    settings: Settings,
    questions: list[Question],
    submissions: list[Submission],
) -> int:
    """Build the Question Grades sheet."""
    headers = [
        "Institutional Student ID",
        "Full Name",
        "Anonymous Code",
    ]
    for q in questions:
        headers.append(f"Q{q.question_number} Grade")
        headers.append(f"Q{q.question_number} Max")
    headers.extend(["Final Grade", "Maximum Grade"])

    ws.append(headers)
    _style_header(ws, len(headers))

    row_num = 1
    for sub in submissions:
        row_num += 1
        anon = sub.anonymous_student
        if anon is None:
            continue

        try:
            ei = get_export_identity(session, anon.id, settings)
        except Exception:
            ei = None

        grade_map = _get_grade_map(session, sub.id)
        qg_vals: list[Decimal] = [g.grade for g in grade_map.values() if g.grade is not None]
        total_g = sum(qg_vals, Decimal("0"))

        ws.cell(row=row_num, column=1, value=_safe_value(ei.institutional_student_id if ei else None))
        ws.cell(row=row_num, column=2, value=_safe_value(ei.full_name if ei else ""))
        ws.cell(row=row_num, column=3, value=anon.anonymous_code)

        col = 4
        for q in questions:
            gr = grade_map.get(q.id)
            grade_val = float(gr.grade) if gr and gr.grade is not None else None
            max_val = float(q.maximum_grade)
            ws.cell(row=row_num, column=col, value=grade_val)
            ws.cell(row=row_num, column=col + 1, value=max_val)
            col += 2

        ws.cell(row=row_num, column=col, value=float(total_g))
        ws.cell(row=row_num, column=col + 1, value=float(assessment.maximum_grade))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num}"
    _auto_width(ws, len(headers))
    return row_num - 1


def _build_feedback_sheet(
    ws: Any,  # noqa: ANN401
    session: Session,
    assessment: Assessment,
    settings: Settings,
    questions: list[Question],
    submissions: list[Submission],
) -> int:
    """Build the Feedback sheet."""
    headers = [
        "Institutional Student ID",
        "Full Name",
        "Anonymous Code",
        "Question Number",
        "Grade",
        "Maximum Grade",
        "Feedback",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    row_num = 1
    for sub in submissions:
        anon = sub.anonymous_student
        if anon is None:
            continue

        try:
            ei = get_export_identity(session, anon.id, settings)
        except Exception:
            ei = None

        grade_map = _get_grade_map(session, sub.id)

        for q in questions:
            row_num += 1
            gr = grade_map.get(q.id)
            ws.cell(row=row_num, column=1, value=_safe_value(ei.institutional_student_id if ei else None))
            ws.cell(row=row_num, column=2, value=_safe_value(ei.full_name if ei else ""))
            ws.cell(row=row_num, column=3, value=anon.anonymous_code)
            ws.cell(row=row_num, column=4, value=q.question_number)
            grade_val = float(gr.grade) if gr and gr.grade is not None else None
            ws.cell(row=row_num, column=5, value=grade_val)
            ws.cell(row=row_num, column=6, value=float(q.maximum_grade))
            ws.cell(row=row_num, column=7, value=_safe_value(gr.feedback if gr else ""))

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num}"
    _auto_width(ws, len(headers))
    return row_num - 1


def _build_export_summary_sheet(
    ws: Any,  # noqa: ANN401
    session: Session,
    assessment: Assessment,
    questions: list[Question],
    submissions: list[Submission],
    export_ref: str,
) -> None:
    """Build the Export Summary sheet."""
    now = datetime.now(UTC)

    grades_list: list[float] = []
    for sub in submissions:
        if sub.anonymous_student is None:
            continue
        grade_map = _get_grade_map(session, sub.id)
        fb_vals: list[Decimal] = [g.grade for g in grade_map.values() if g.grade is not None]
        total_g = sum(fb_vals, Decimal("0"))
        grades_list.append(float(total_g))

    approved_count = sum(1 for s in submissions if s.review_status == "approved")
    avg = sum(grades_list) / max(len(grades_list), 1) if grades_list else 0.0
    min_g = min(grades_list) if grades_list else 0.0
    max_g = max(grades_list) if grades_list else 0.0

    summary_data = [
        ("Export Reference", export_ref),
        ("Assessment Title", assessment.title),
        ("Material", assessment.material.name if assessment.material else ""),
        ("Academic Year", assessment.academic_year or ""),
        ("Assessment Maximum", float(assessment.maximum_grade)),
        ("Finalization Timestamp", str(assessment.finalized_at or "")),
        ("Export Timestamp", str(now)),
        ("Number of Submissions", len(submissions)),
        ("Approved Count", approved_count),
        ("Average Grade", round(avg, 2)),
        ("Minimum Grade", round(min_g, 2)),
        ("Maximum Grade", round(max_g, 2)),
        ("Workbook Schema Version", SCHEMA_VERSION),
    ]

    for key, value in summary_data:
        ws.append([key, value])

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30


def generate_export_workbook(
    session: Session,
    assessment_id: str,
    settings: Settings,
) -> ExportResult:
    """Generate an Excel workbook for a finalized assessment.

    Parameters
    ----------
    session : Session
        Database session.
    assessment_id : str
        ID of the finalized assessment.
    settings : Settings
        Application settings (for encryption key).

    Returns
    -------
    ExportResult
        Workbook bytes, hash, size, row count, and reference.

    Raises
    ------
    FinalizedAssessmentExportError
        If the assessment is not finalized or data is invalid.
    """
    assessment = session.query(Assessment).filter(Assessment.id == assessment_id).first()
    if assessment is None:
        raise FinalizedAssessmentExportError(f"Assessment {assessment_id[:8]} not found.")
    if assessment.finalization_status != "finalized":
        raise FinalizedAssessmentExportError(
            f"Assessment {assessment_id[:8]} is not finalized (status={assessment.finalization_status})."
        )

    questions = _get_questions(session, assessment_id)
    if not questions:
        raise FinalizedAssessmentExportError("Assessment has no questions.")

    submissions = _get_submissions(session, assessment_id)
    if not submissions:
        raise FinalizedAssessmentExportError("Assessment has no submissions.")

    # Validate data integrity
    for sub in submissions:
        if sub.anonymous_student is None:
            raise ExportValidationError(f"Submission {sub.id[:8]} has no anonymous student.")
        grade_map = _get_grade_map(session, sub.id)
        for q in questions:
            gr = grade_map.get(q.id)
            if gr is None:
                raise ExportValidationError(
                    f"Submission {sub.id[:8]} Q{q.question_number}: missing GradeRecord."
                )
            if gr.grade is None:
                raise ExportValidationError(
                    f"Submission {sub.id[:8]} Q{q.question_number}: null grade."
                )

    export_ref = f"EXP-{uuid.uuid4().hex[:8].upper()}"

    try:
        wb = Workbook()

        # Sheet 1: Final Grades
        ws1 = wb.active
        if ws1 is not None:
            ws1.title = "Final Grades"
        row_count = _build_final_grades_sheet(ws1, session, assessment, settings, questions, submissions)

        # Sheet 2: Question Grades
        ws2 = wb.create_sheet("Question Grades")
        _build_question_grades_sheet(ws2, session, assessment, settings, questions, submissions)

        # Sheet 3: Feedback
        ws3 = wb.create_sheet("Feedback")
        _build_feedback_sheet(ws3, session, assessment, settings, questions, submissions)

        # Sheet 4: Export Summary
        ws4 = wb.create_sheet("Export Summary")
        _build_export_summary_sheet(ws4, session, assessment, questions, submissions, export_ref)

        # Workbook metadata
        wb.properties.creator = "Academic Anonymous Grader"
        wb.properties.title = f"Final Grades - {assessment.title}"
        wb.properties.description = f"Exported by Academic Anonymous Grader v{SCHEMA_VERSION}"

        # Write to bytes
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        workbook_bytes = buf.read()

        sha256 = hashlib.sha256(workbook_bytes).hexdigest()

        return ExportResult(
            workbook_bytes=workbook_bytes,
            file_hash=sha256,
            file_size=len(workbook_bytes),
            row_count=row_count,
            export_reference=export_ref,
        )
    except Exception as exc:
        raise ExportWorkbookError(f"Workbook generation failed: {exc}") from exc
